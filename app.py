# 1- Entrar na planilha e extrair o cpf do cliente
# 2- Entrar no site https://consultcpf-devaprender.netlify.app/ consultar o cpf da planilha para pesquisar o status do pgamento daquele cliente
# 3- Verificar se esta "em dia" ou "atrasado"
# 4- Se esta "em dia" , pegar a data do pagamento e o metodo de pagamento
# 5- Caso contrario(se estiver atrasado), colocar o status como pendente
# 6- inserir essas informacoes em uma nova planilha

import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep

planilha_clientes = openpyxl.load_workbook("dados_clientes.xlsx")
pagina_clientes = planilha_clientes['Sheet1']

site_consulta = "https://consultcpf-devaprender.netlify.app/"
driver = webdriver.Chrome()

for linha in pagina_clientes.iter_rows(min_row=2, values_only=True):
    nome, valor, cpf, vencimento = linha

    driver.get(site_consulta)
    sleep(5)

    campo_pesquisa = driver.find_element(By.XPATH, '//input[@id="cpfInput"]')
    sleep(1)
    campo_pesquisa.clear()
    campo_pesquisa.send_keys(cpf)
    sleep(1)

    botao_consulta = driver.find_element(By.XPATH, '//button[@class ="btn btn-custom btn-lg btn-block mt-3"]')
    sleep(1)
    botao_consulta.click()

    sleep(4)
    status = driver.find_element(By.XPATH, '//span[@id="statusLabel"]')

    if status.text == "em dia":
        data_pagamento = driver.find_element(By.XPATH, '//p[@id="paymentDate"]')
        metodo_pagamento = driver.find_element(By.XPATH, '//p[@id="paymentMethod"]')

        data_pagamento_limpo = data_pagamento.text.split()[3]
        metodo_pagamento_limpo = metodo_pagamento.text.split()[3]

        planilha_fechamento = openpyxl.load_workbook("planilha fechamento.xlsx")
        pagina_fechamento = planilha_fechamento["Sheet1"]

        pagina_fechamento.append([nome, valor, cpf, vencimento, 'em dia', data_pagamento_limpo, metodo_pagamento_limpo])
        planilha_fechamento.save("planilha fechamento.xlsx")
    else:
        planilha_fechamento = openpyxl.load_workbook("planilha fechamento.xlsx")
        pagina_fechamento = planilha_fechamento["Sheet1"]

        pagina_fechamento.append([nome, valor, cpf, vencimento, 'pendente'])
        planilha_fechamento.save("planilha fechamento.xlsx")

